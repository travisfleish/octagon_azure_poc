#!/usr/bin/env python3
"""
SOW to Excel Mapper - Creates comprehensive Excel spreadsheet of all extracted fields
=================================================================================

This script creates an Excel spreadsheet showing the complete mapping of every
standardized field for each SOW document.
"""

import json
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def load_analysis_results():
    """Load the SOW analysis results"""
    with open('real_sow_analysis_results.json', 'r') as f:
        return json.load(f)

def create_comprehensive_excel(results):
    """Create a comprehensive Excel file with all SOW data"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create main summary sheet
    create_summary_sheet(wb, results)
    
    # Create detailed SOW data sheet
    create_detailed_sow_sheet(wb, results)
    
    # Create departments breakdown sheet
    create_departments_sheet(wb, results)
    
    # Create deliverables analysis sheet
    create_deliverables_sheet(wb, results)
    
    # Create roles analysis sheet
    create_roles_sheet(wb, results)
    
    # Create company analysis sheet
    create_company_sheet(wb, results)
    
    # Save the workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"octagon_sow_complete_mapping_{timestamp}.xlsx"
    wb.save(filename)
    
    return filename

def create_summary_sheet(wb, results):
    """Create the main summary sheet"""
    ws = wb.create_sheet("Summary", 0)
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add title
    ws['A1'] = "Octagon SOW Analysis - Complete Field Mapping"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws.merge_cells('A1:H1')
    
    # Summary statistics
    summary_data = [
        ["Analysis Summary", ""],
        ["Total SOWs Analyzed", results['analysis_summary']['total_sows_analyzed']],
        ["Successful Extractions", results['analysis_summary']['successful_extractions']],
        ["Success Rate", f"{(results['analysis_summary']['successful_extractions']/results['analysis_summary']['total_sows_analyzed']*100):.1f}%"],
        ["Average Confidence", f"{results['analysis_summary']['average_confidence']:.2f}"],
        ["Analysis Timestamp", results['analysis_summary']['analysis_timestamp']],
        ["", ""],
        ["Data Summary", ""],
        ["Companies Identified", len(results['companies'])],
        ["Departments Mapped", len(results['departments'])],
        ["Unique Deliverables", len(results['deliverables'])],
        ["Unique Roles", len(results['roles'])],
        ["Duration Range", f"{results['duration_analysis']['min_weeks']}-{results['duration_analysis']['max_weeks']} weeks"],
        ["Average Duration", f"{results['duration_analysis']['avg_weeks']:.1f} weeks"],
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, start=3):
        ws[f'A{row_idx}'] = label
        ws[f'B{row_idx}'] = value
        if label in ["Analysis Summary", "Data Summary", ""]:
            ws[f'A{row_idx}'].font = Font(bold=True)
    
    # Department breakdown
    ws['A20'] = "Department Usage Across SOWs"
    ws['A20'].font = Font(bold=True, size=12)
    
    dept_data = [["Department", "Count", "Percentage"]]
    total_sows = results['analysis_summary']['successful_extractions']
    for dept, count in results['departments'].items():
        percentage = (count / total_sows) * 100
        dept_data.append([dept, count, f"{percentage:.1f}%"])
    
    for row_idx, row_data in enumerate(dept_data, start=21):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if row_idx == 21:  # Header row
                cell.font = header_font
                cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if hasattr(cell, 'column_letter') and column_letter is None:
                column_letter = cell.column_letter
            try:
                if hasattr(cell, 'value') and cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

def create_detailed_sow_sheet(wb, results):
    """Create detailed SOW data sheet"""
    ws = wb.create_sheet("Detailed SOW Data")
    
    # Get all successful results
    successful_results = [r for r in results['detailed_results'] if r['confidence_score'] > 0]
    
    # Prepare data for DataFrame
    sow_data = []
    for result in successful_results:
        # Flatten the data
        row = {
            'File Name': result['file_name'],
            'Company': result['company'],
            'Project Title': result['project_title'],
            'Duration (Weeks)': result['duration_weeks'],
            'Departments (Count)': len(result['departments_involved']),
            'Departments List': ', '.join(result['departments_involved']),
            'Deliverables (Count)': len(result['deliverables']),
            'Roles (Count)': len(result['roles_mentioned']),
            'Confidence Score': result['confidence_score'],
            'Scope Description': result['scope_description'][:200] + '...' if len(result['scope_description']) > 200 else result['scope_description']
        }
        
        # Add budget info if available
        budget_info = result.get('budget_info', {})
        row['Budget Total'] = budget_info.get('total_budget', 'N/A')
        row['Budget Currency'] = budget_info.get('budget_currency', 'N/A')
        
        sow_data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(sow_data)
    
    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if hasattr(cell, 'column_letter') and column_letter is None:
                column_letter = cell.column_letter
            try:
                if hasattr(cell, 'value') and cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

def create_departments_sheet(wb, results):
    """Create departments breakdown sheet"""
    ws = wb.create_sheet("Departments Analysis")
    
    # Department mapping data
    dept_mapping = {
        'client_services': 'Client Services (Account Management, Client Relationships)',
        'strategy': 'Strategy (Strategic Planning, Insights, Analytics)',
        'planning_creative': 'Planning & Creative (Creative Development, Brand Work, Campaign Planning)',
        'integrated_production_experiences': 'Integrated Production & Experiences (Events, Hospitality, Activations, Production)'
    }
    
    # Create department analysis
    dept_data = []
    for dept, count in results['departments'].items():
        dept_data.append({
            'Department Code': dept,
            'Department Description': dept_mapping.get(dept, dept),
            'SOW Count': count,
            'Percentage': f"{(count / results['analysis_summary']['successful_extractions']) * 100:.1f}%"
        })
    
    # Add department usage by SOW
    ws.append(['Department Code', 'Department Description', 'SOW Count', 'Percentage'])
    
    for row in dept_data:
        ws.append([row['Department Code'], row['Department Description'], row['SOW Count'], row['Percentage']])
    
    # Add detailed breakdown by SOW
    ws.append([])
    ws.append(['SOW File', 'Company', 'Departments Involved'])
    
    successful_results = [r for r in results['detailed_results'] if r['confidence_score'] > 0]
    for result in successful_results:
        ws.append([
            result['file_name'],
            result['company'],
            ', '.join(result['departments_involved'])
        ])
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for row_num in [1, 6]:  # Header rows
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if hasattr(cell, 'column_letter') and column_letter is None:
                column_letter = cell.column_letter
            try:
                if hasattr(cell, 'value') and cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

def create_deliverables_sheet(wb, results):
    """Create deliverables analysis sheet"""
    ws = wb.create_sheet("Deliverables Analysis")
    
    # Get all deliverables with counts
    deliverables = results['deliverables']
    
    # Sort by frequency
    sorted_deliverables = sorted(deliverables.items(), key=lambda x: x[1], reverse=True)
    
    # Add headers
    ws.append(['Deliverable', 'Frequency', 'SOWs'])
    
    # Add deliverables data
    for deliverable, count in sorted_deliverables:
        # Find which SOWs contain this deliverable
        sow_list = []
        successful_results = [r for r in results['detailed_results'] if r['confidence_score'] > 0]
        for result in successful_results:
            if any(deliverable.lower() in d.lower() for d in result['deliverables']):
                sow_list.append(result['file_name'])
        
        ws.append([
            deliverable[:100] + '...' if len(deliverable) > 100 else deliverable,
            count,
            ', '.join(sow_list[:3]) + ('...' if len(sow_list) > 3 else '')
        ])
    
    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_roles_sheet(wb, results):
    """Create roles analysis sheet"""
    ws = wb.create_sheet("Roles Analysis")
    
    # Get all roles with counts
    roles = results['roles']
    
    # Sort by frequency
    sorted_roles = sorted(roles.items(), key=lambda x: x[1], reverse=True)
    
    # Add headers
    ws.append(['Role', 'Frequency', 'SOWs'])
    
    # Add roles data
    for role, count in sorted_roles:
        # Find which SOWs contain this role
        sow_list = []
        successful_results = [r for r in results['detailed_results'] if r['confidence_score'] > 0]
        for result in successful_results:
            if any(role.lower() in r.lower() for r in result['roles_mentioned']):
                sow_list.append(result['file_name'])
        
        ws.append([
            role,
            count,
            ', '.join(sow_list[:3]) + ('...' if len(sow_list) > 3 else '')
        ])
    
    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if hasattr(cell, 'column_letter') and column_letter is None:
                column_letter = cell.column_letter
            try:
                if hasattr(cell, 'value') and cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

def create_company_sheet(wb, results):
    """Create company analysis sheet"""
    ws = wb.create_sheet("Company Analysis")
    
    # Company data
    companies = results['companies']
    
    # Add headers
    ws.append(['Company', 'SOW Count', 'Total Duration (Weeks)', 'Average Duration (Weeks)', 'Projects'])
    
    # Add company data
    for company, data in companies.items():
        if company != 'ERROR':
            avg_duration = data['total_duration_weeks'] / data['sow_count'] if data['sow_count'] > 0 else 0
            ws.append([
                company,
                data['sow_count'],
                data['total_duration_weeks'],
                f"{avg_duration:.1f}",
                '; '.join(data['projects'])
            ])
    
    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

def main():
    """Main function to create the Excel file"""
    
    print("📊 Creating Comprehensive Excel Spreadsheet...")
    
    # Load results
    results = load_analysis_results()
    
    # Create Excel file
    filename = create_comprehensive_excel(results)
    
    print(f"\n✅ Excel file created successfully: {filename}")
    print(f"\n📋 The spreadsheet contains 6 sheets:")
    print("   1. Summary - Overview and statistics")
    print("   2. Detailed SOW Data - Complete field mapping for each SOW")
    print("   3. Departments Analysis - Department usage and mapping")
    print("   4. Deliverables Analysis - All deliverables with frequency")
    print("   5. Roles Analysis - All roles with frequency")
    print("   6. Company Analysis - Company breakdown and projects")
    
    print(f"\n🎯 This gives you the complete mapping of every standardized field for each SOW!")

if __name__ == "__main__":
    main()
